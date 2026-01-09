from typing import Optional
from uuid import UUID
from fastapi import Depends
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break
from sqlalchemy import select
from app.db import get_db
from sqlalchemy.orm import Session
from app.exceptions import CustomException
from app.models.models import Company, Product, Shop


def export_company(
    company_id: UUID,
    db: Session = Depends(get_db),
):
    export_company_service(company_id, db=db)


def export_shop(shop_id: UUID, db: Session = Depends(get_db)):
    stmt = select(Shop).join(Company).where(Shop.is_active == True, Shop.id == shop_id)
    shop = db.execute(stmt).scalar_one_or_none()
    if not shop:
        raise CustomException(status_code=422, detail={"message": "Shop not found"})

    workbook = Workbook()

    for company in shop.companies:
        export_company_service  (company.id, existing_workbook=workbook, db=db)

    workbook.save(f"{shop.name.upper()}.xlsx")


def export_company_service(
    company_id: UUID,
    db: Session = Depends(get_db),
    existing_workbook: Optional[Workbook] = None,
):
    stmt = select(Company).where(Company.is_active == True, Company.id == company_id)
    company = db.execute(stmt).scalar_one_or_none()
    if not company:
        raise CustomException(status_code=422, detail={"message": "Company not found"})

    products: list[Product] = company.product

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    if not existing_workbook:
        workbook = Workbook()
        workbook.remove(workbook["Sheet"])

    workbook.create_sheet(title=company.name)
    worksheet = workbook[company.name]

    title_font = Font(name="Arial", size=24, bold=True, color="000000")
    product_font = Font(size=14, bold=True, color="000000")

    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = company.name.upper()
    worksheet["A1"].font = title_font
    worksheet["A1"].border = thin_border

    center_alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].alignment = center_alignment

    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_margins.bottom = 0.5
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.left = 0.75
    worksheet.page_margins.right = 0.5

    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = 1

    worksheet.row_dimensions[1].height = 40

    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 56

    row: int = 2
    current_height: int = 40

    for product in products:
        if (current_height + (product.rows * 30)) > 750:
            current_height = 0
            worksheet.row_breaks.append(Break(id=row - 1))

        current_height += product.rows * 30

        for i in range(product.rows):
            worksheet.row_dimensions[row + i].height = 30
            worksheet[f"C{row + i}"].border = thin_border

        worksheet[f"A{row}"] = product.name.upper()
        worksheet[f"B{row}"] = "₹" + str(round(product.price))

        worksheet["A" + str(row)].font = product_font
        worksheet["B" + str(row)].font = product_font

        worksheet["A" + str(row)].alignment = center_alignment
        worksheet["B" + str(row)].alignment = center_alignment

        worksheet["A" + str(row)].border = thin_border
        worksheet["B" + str(row)].border = thin_border

        worksheet.merge_cells(f"A{row}:A{row + product.rows - 1}")
        worksheet.merge_cells(f"B{row}:B{row + product.rows - 1}")
        row += product.rows

    if not existing_workbook:
        workbook.save(f"{company.name.upper()}.xlsx")
